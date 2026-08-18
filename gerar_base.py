#!/usr/bin/env python3
"""Gera a planilha Base (Payments <data>.xlsx) a partir dos dois
relatórios brutos do SAP — sem depender de Power Query.

Uso:
    python gerar_base.py
    python gerar_base.py --file1 <caminho.xlsx> --file2 <caminho.xlsx> [--output-dir <pasta>]

Sem --file1/--file2, lê os dois arquivos de `entrada/` pelo FORMATO (número de colunas e
cabeçalho), não pelo nome do arquivo — já teve semana que vieram com o
nome trocado. Precisa dos dois:

  - "Administrar itens de fornecedor": 57 colunas, cabeçalho começa com
    'Conta do Razão' na 2ª posição. É a linha de baixa bancária —
    fonte principal da classificação.
  - "Partidas individuais no Razão": 22 colunas, cabeçalho começa com
    'Bloqueio pgto.item' na 2ª posição. Usado só pra achar a referência
    da fatura de cada lançamento ICBRWA (regra PEN).

Exceções manuais da semana (tudo que é PIX — não tem como detectar
automaticamente) vão em `OVERRIDES_SEMANA` abaixo. Editar toda semana.

Saída: saida/<AAAA-MM-DD>/Payments_<data>.xlsx, aba SAP (linha a linha,
classificada) + aba Resumo (pivot por produto, igual ao layout que vai
no corpo do e-mail de aprovação).
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
import regras

# ============================================================
# EDITAR TODA SEMANA — exceções de PIX que não têm padrão detectável.
# Chave: número do documento de pagamento (Lançamento contábil / Business
# Transaction Type). Valor: sempre "PIX" ou "Pagamento Fornecedores".
# Ver docstring de regras.classificar() pros casos já documentados.
# ============================================================
OVERRIDES_SEMANA: dict[str, str] = {
    "1300005654": "PIX",                     # BESSA SERVICOS — PIX avulso
    "1300005682": "Pagamento Fornecedores",  # Folha paga via PIX (sigilo)
    # Exemplo de exclusão de duplicidade (semana de 15.07.2026):
    # "1300005522": regras.EXCLUIDO,  # duplicava pagamento já em TM5
    # "1300005270": "Pagamento Fornecedores",  # ICBRWA sem PEN, não é Folha
}


def _identificar(caminho: Path) -> tuple[str, pd.DataFrame]:
    """Lê um export SAP e devolve (tipo, dataframe) identificando pelas
    colunas essenciais presentes — não pela posição, nem pela quantidade
    total de colunas (o SAP já mandou o mesmo relatório com 57 e com 13
    colunas, mudando só quais campos vêm marcados no export)."""
    df = pd.read_excel(caminho, sheet_name="Exportação SAPUI5")
    cols = set(df.columns)
    essenciais_banco = {"Conta do Razão", "Lançamento contábil", "Tipo lçto.contábil", "Mont.moeda empresa"}
    essenciais_fornecedor = {"Bloqueio pgto.item", "Fornecedor", "Lançto.compensação"}
    if essenciais_banco <= cols and "Bloqueio pgto.item" not in cols:
        return "banco", df
    if essenciais_fornecedor <= cols:
        return "partidas", df
    raise ValueError(
        f"{caminho.name}: formato não reconhecido ({len(cols)} colunas: "
        f"{sorted(cols)[:6]}...). Confirmar se o export do SAP não mudou de layout."
    )


def _reconstituir_contrapartida(df_banco: pd.DataFrame, df_partidas: pd.DataFrame) -> pd.DataFrame:
    """Se o export de 'Partidas individuais no Razão' veio sem
    'Cta.contrapartida' (aconteceu na semana de 22.07.2026 — SAP mandou
    versão reduzida com 13 colunas), reconstitui essa coluna cruzando
    pelo Lançto.compensação com o Fornecedor do relatório 'Administrar
    itens de fornecedor'. Sem isso as regras ICBRWA/Caixa não funcionam,
    porque dependem do texto dessa coluna."""
    if "Cta.contrapartida" in df_banco.columns:
        return df_banco

    doc_pag = regras.CAMPOS_PARTIDAS["lancto_compensacao"]
    fornecedor = regras.CAMPOS_PARTIDAS["fornecedor"]
    nome = "Nome do fornecedor"

    lookup: dict[str, str] = {}
    for _, linha in df_partidas.iterrows():
        doc = regras.doc_str(linha.get(doc_pag, ""))
        forn = str(linha.get(fornecedor, "") or "")
        nom = str(linha.get(nome, "") or "")
        if doc and doc != "None" and forn:
            lookup[doc] = f"{forn} ({nom})"

    chave = df_banco["Lançamento contábil"].map(regras.doc_str)
    df_banco = df_banco.copy()
    df_banco["Cta.contrapartida"] = chave.map(lookup).fillna("")
    return df_banco


def _montar_par(achados: dict[str, pd.DataFrame], onde: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Confere que os dois formatos (banco/partidas) foram achados em
    `achados` e devolve o par pronto pra classificar_semana()."""
    faltando = {"banco", "partidas"} - achados.keys()
    if faltando:
        raise FileNotFoundError(
            f"Faltam relatório(s) {onde}: {sorted(faltando)}. "
            "Precisa dos dois: 'Administrar itens de fornecedor' e "
            "'Partidas individuais no Razão'."
        )
    df_banco = _reconstituir_contrapartida(achados["banco"], achados["partidas"])
    return df_banco, achados["partidas"]


def carregar_entrada(dir_entrada: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Acha os dois relatórios em `dir_entrada` pelo formato, não pelo nome."""
    candidatos = [p for p in dir_entrada.glob("*.xlsx") if not p.name.startswith("~$")]
    achados: dict[str, pd.DataFrame] = {}
    for caminho in candidatos:
        tipo, df = _identificar(caminho)
        achados[tipo] = df
    return _montar_par(achados, f"em {dir_entrada}/")


def carregar_entrada_arquivos(caminhos: list[Path]) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Mesma identificação por formato de `carregar_entrada()`, mas para uma
    lista explícita de arquivos (ex.: os dois selecionados pelo usuário na
    UI) em vez de escanear um diretório."""
    achados: dict[str, pd.DataFrame] = {}
    for caminho in caminhos:
        tipo, df = _identificar(caminho)
        achados[tipo] = df
    return _montar_par(achados, "nos arquivos selecionados")


def _vendor(cta_contrapartida: object) -> str:
    """'105984 (TRADEWORKS SERVICOS...)' -> 'TRADEWORKS SERVICOS...'."""
    texto = str(cta_contrapartida or "").strip()
    if "(" in texto:
        return texto.split("(", 1)[1].rstrip(")").strip()
    return texto


NOME_BANCO = {
    "1001002000": "Santander",
    "1000000050": "Banco do Brasil",
}


def _banco(conta_razao: object) -> str:
    codigo = str(conta_razao or "").split()[0] if str(conta_razao or "").split() else ""
    return NOME_BANCO.get(codigo, codigo)


def classificar_semana(df_banco: pd.DataFrame, df_partidas: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    refs = regras.referencias_fatura_icbrwa(df_partidas)
    textos = regras.textos_item_por_documento(df_partidas)
    out = regras.classificar(df_banco, overrides=OVERRIDES_SEMANA, referencias_fatura=refs, textos_item=textos)
    out = out.assign(Banco=out[regras.CAMPOS_BRUTOS["conta_razao"]].map(_banco))

    c = regras.CAMPOS_BRUTOS
    limpo = pd.DataFrame({
        "Company Code": out["Empresa"] if "Empresa" in out.columns else "4690",
        "G/L Account": out[c["conta_razao"]],
        "Documento": out[c["documento"]].map(regras.doc_str),
        "Tipo Lançamento": out[c["tipo_lancamento"]],
        "Vendor": out[c["contrapartida"]].map(_vendor),
        "Produto": out["Produto"],
        "Banco": out["Banco"],
        "Status": "Pendentes",
        "Posting Date": pd.to_datetime(out[c["data"]]).dt.date,
        "Valor": out[c["valor"]].round(2),
        "RegraAplicada": out["RegraAplicada"],
        # True se o documento de pagamento tem linha de fatura correspondente
        # no export "Administrar itens de fornecedor". Quando é False e a
        # regra aplicada foi R4 (ICBRWA -> Folha), a classificação é um
        # palpite do padrão, não evidência — ver check na aba Validações.
        "TemFatura": out[c["documento"]].map(regras.doc_str).isin(textos.keys()),
    }).sort_values(["Produto", "Valor"]).reset_index(drop=True)

    # Base = o relatório "Administrar itens de fornecedor" original (22
    # colunas, o mesmo que você conferiu no print), com "Produto" e "Banco"
    # anexados via join por Lançto.compensação = documento de pagamento
    # (o mesmo que alimenta a classificação). Fica em branco pra fatura
    # que não tem pagamento nesta janela (ex.: ainda em aberto).
    lookup = out.set_index(out[c["documento"]].map(regras.doc_str))[["Produto", "Banco"]]
    lookup = lookup[~lookup.index.duplicated(keep="first")]

    completa = df_partidas.copy()
    comp_col = regras.CAMPOS_PARTIDAS["lancto_compensacao"]
    chave = completa[comp_col].map(regras.doc_str)
    completa["Produto"] = chave.map(lookup["Produto"])
    completa["Banco"] = chave.map(lookup["Banco"])
    completa = completa.sort_values(
        ["Produto", regras.CAMPOS_PARTIDAS["referencia"]], na_position="last"
    ).reset_index(drop=True)

    return limpo, completa


# ---------------- escrita do xlsx ----------------

FMT = "#,##0.00;-#,##0.00;-"
PCT = "0.0%"
ARIAL = "Arial"
H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(name=ARIAL, size=9, bold=True, color="FFFFFF")
T_FILL = PatternFill("solid", fgColor="D9E1F2")
BODY = Font(name=ARIAL, size=9)
BOLD = Font(name=ARIAL, size=9, bold=True)
TITLE = Font(name=ARIAL, size=9, bold=True, color="1F3864")
LABEL = Font(name=ARIAL, size=9, bold=True, color="404040")
NOTE = Font(name=ARIAL, size=9, italic=True, color="808080")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_BLOQUEIO = PatternFill("solid", fgColor="FFC7CE")
FILL_ALERTA = PatternFill("solid", fgColor="FFEB9C")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
_FILL_POR_SEV = {"BLOQUEIO": FILL_BLOQUEIO, "ALERTA": FILL_ALERTA, "OK": FILL_OK}

# Cores de banco — usadas só nos cabeçalhos das colunas Santander/BB da
# "Dinâmica por produto". Santander = vermelho da marca; BB = amarelo da
# marca, com texto no azul da marca (amarelo puro não lê bem com branco).
SANTANDER_FILL = PatternFill("solid", fgColor="EC0000")
SANTANDER_FONT = Font(name=ARIAL, size=9, bold=True, color="FFFFFF")
BB_FILL = PatternFill("solid", fgColor="FFCC00")
BB_FONT = Font(name=ARIAL, size=9, bold=True, color="003087")


def _cabecalho(ws, row, cols, widths, col_inicio=1):
    for offset, (nome, w) in enumerate(zip(cols, widths)):
        i = col_inicio + offset
        cell = ws.cell(row=row, column=i, value=nome)
        cell.fill, cell.font, cell.border = H_FILL, H_FONT, BOX
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w


def _linha_total(ws, row, ncols, rotulo_col, rotulo):
    ws.cell(row=row, column=rotulo_col, value=rotulo)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill, cell.border = BOLD, T_FILL, BOX


def _checks(base: pd.DataFrame, janela: "config.Janela") -> list[tuple[str, str, str]]:
    """(check, severidade, detalhe) — espelha validacoes.py, adaptado pra
    ler direto do `limpo` (Company Code/Documento/Data etc. em vez do
    fluxo antigo baseado em Payments <data>.xlsx já classificado)."""
    checks: list[tuple[str, str, str]] = []

    empresas = sorted(base["Company Code"].dropna().astype(str).unique())
    if len(empresas) == 1:
        checks.append(("Company Code único", "OK", empresas[0]))
    else:
        checks.append(("Company Code único", "BLOQUEIO", f"achado {empresas}"))

    contas = sorted(base["G/L Account"].dropna().astype(str).unique())
    contas_validas = {c for c in contas if regras.conta_banco_valida(c)}
    if contas and contas_validas == set(contas):
        checks.append(("Conta G/L única", "OK", ", ".join(contas)))
    else:
        checks.append(("Conta G/L única", "BLOQUEIO", f"achado {contas}"))

    status = sorted(base["Status"].dropna().unique())
    if status == ["Pendentes"]:
        checks.append(("Status único", "OK", "Pendentes"))
    else:
        checks.append(("Status único", "BLOQUEIO", f"achado {status}"))

    datas = sorted({pd.Timestamp(d).date() for d in base["Posting Date"].dropna()})
    if datas == [janela.data_referencia]:
        checks.append(("Posting Date == Data_Referencia", "OK", janela.data_referencia.strftime("%d/%m/%Y")))
    else:
        checks.append(("Posting Date == Data_Referencia", "BLOQUEIO",
                        f"esperado {janela.data_referencia:%d/%m/%Y}, achado {[d.strftime('%d/%m/%Y') for d in datas]}"))

    dup = base.loc[base["Documento"].duplicated(keep=False), "Documento"]
    if dup.empty:
        checks.append(("Documentos únicos", "OK", f"{len(base)} documentos"))
    else:
        checks.append(("Documentos únicos", "BLOQUEIO", f"{dup.nunique()} duplicado(s)"))

    pos = base[base["Valor"] > 0]
    if pos.empty:
        checks.append(("Todos os valores são saída", "OK", "nenhum valor positivo"))
    else:
        checks.append(("Todos os valores são saída", "BLOQUEIO", f"{len(pos)} lançamento(s) positivo(s)"))

    zero = base[base["Valor"] == 0]
    if zero.empty:
        checks.append(("Sem lançamentos zerados", "OK", ""))
    else:
        checks.append(("Sem lançamentos zerados", "ALERTA", f"{len(zero)} lançamento(s) com valor 0"))

    produtos_validos = {"Folha de Pagamento", "PIX", "Pagamento Fornecedores", "TM5"}
    nulos = base["Produto"].isna().sum()
    invalidos = sorted(set(base["Produto"].dropna()) - produtos_validos)
    if nulos or invalidos:
        checks.append(("Produto válido em todas as linhas", "BLOQUEIO", f"{nulos} sem produto; fora da lista: {invalidos}"))
    else:
        checks.append(("Produto válido em todas as linhas", "OK", f"{base['Produto'].nunique()} produtos"))

    nulos_v = base["Vendor"].isna().sum()
    if nulos_v:
        checks.append(("Fornecedor preenchido", "BLOQUEIO", f"{nulos_v} linha(s) sem fornecedor"))
    else:
        checks.append(("Fornecedor preenchido", "OK", f"{base['Vendor'].nunique()} fornecedores"))

    # Um ICBRWA sem linha de fatura no export cai em Folha só pelo padrão R4,
    # sem nenhuma evidência de texto. Aconteceu em 19.08.2026: o export veio
    # sem a fatura 5015173 (doc 1300006133, -93.551,92) e a classificação
    # ficou sendo um palpite — acertou, mas às cegas. Sinalizar pra conferir.
    # Cobre os DOIS produtos que dependem do par (código do fornecedor +
    # mensagem no "Texto de item"): R4 = ICBRWA -> Folha, R5 = Caixa -> PIX.
    # Quando a fatura não veio no export não há texto pra confirmar, e a
    # classificação sai só pelo código do fornecedor. Não é erro — é o único
    # dado disponível — mas fica sinalizado pra conferência manual.
    if "TemFatura" in base.columns and "RegraAplicada" in base.columns:
        regra = base["RegraAplicada"].astype(str)
        so_por_fornecedor = regra.str.startswith("R4 —") | regra.str.startswith("R5 —")
        cegos = base[so_por_fornecedor & ~base["TemFatura"]]
        if cegos.empty:
            checks.append((
                "Folha/PIX confirmados por texto de item", "OK",
                "todo Folha e PIX tem código do fornecedor + mensagem no texto",
            ))
        else:
            detalhe = "; ".join(
                f"{r.Documento} ({r.Produto}) {r.Valor:,.2f}"
                for r in cegos.head(5).itertuples(index=False)
            )
            checks.append((
                "Folha/PIX confirmados por texto de item", "ALERTA",
                f"{len(cegos)} doc(s) sem fatura no export — classificado(s) só pelo "
                f"código do fornecedor, sem mensagem pra confirmar: {detalhe}",
            ))

    erros_zp = regras.conferir_zp(base, "Tipo Lançamento", "Produto")
    if erros_zp:
        checks.append(("Coerência R1 (ZP -> TM5)", "BLOQUEIO", "; ".join(erros_zp)))
    else:
        checks.append(("Coerência R1 (ZP -> TM5)", "OK", "sem divergência"))

    for produto in ("Folha de Pagamento", "PIX", "Pagamento Fornecedores", "TM5"):
        total = base.loc[base["Produto"] == produto, "Valor"].sum()
        checks.append((f"Reconciliação — {produto}", "OK", f"{total:,.2f}"))
    checks.append(("Reconciliação — Total Geral", "OK", f"{base['Valor'].sum():,.2f}"))

    return checks


def _aba_base(wb: Workbook, base: pd.DataFrame) -> int:
    ws = wb.active
    ws.title = "Base"
    ws.sheet_view.showGridLines = False
    _cabecalho(ws, 1, ["Documento", "Data", "Status", "Produto", "Fornecedor", "Valor (BRL)"],
               [14, 12, 12, 24, 46, 16])
    cols = ["Documento", "Posting Date", "Status", "Produto", "Vendor", "Valor"]
    for r, row in enumerate(base[cols].itertuples(index=False), start=2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font, cell.border = BODY, BOX
            if c == 2:
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = Alignment(horizontal="center")
            if c == 6:
                cell.number_format = FMT
    last = len(base) + 1
    _linha_total(ws, last + 1, 6, 5, "Total Geral")
    tv = ws.cell(row=last + 1, column=6, value=f"=SUM(F2:F{last})")
    tv.number_format = FMT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{last}"
    return last


def _aba_resumo(wb: Workbook, janela: "config.Janela", base: pd.DataFrame, last: int) -> None:
    ws = wb.create_sheet("Resumo", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Relatório de Aprovação de Pagamentos"
    ws["A1"].font = TITLE
    ws["A2"] = f"Nordex Energy Brasil LTDA — Company Code 4690 — {janela.rotulo}"
    ws["A2"].font = Font(name=ARIAL, size=9, color="404040")

    bancos = sorted(base["Banco"].dropna().unique())
    meta = [
        ("Data de pagamento", janela.rotulo),
        ("Banco", ", ".join(bancos)),
        ("Conta G/L", ", ".join(sorted(base["G/L Account"].dropna().astype(str).unique()))),
        ("Status dos lançamentos", "Pendentes"),
        ("Produtos", "Folha de Pagamento, PIX, Pagamento Fornecedores, TM5"),
        ("Documentos", len(base)),
        ("Fonte", "SAP: Administrar itens de fornecedor + Partidas individuais no Razão"),
    ]
    r = 4
    for k, v in meta:
        ws.cell(row=r, column=1, value=k).font = LABEL
        ws.cell(row=r, column=2, value=v).font = BODY
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Resumo por produto").font = Font(name=ARIAL, size=9, bold=True)
    r += 1
    _cabecalho(ws, r, ["Produto", "Qtd. documentos", "Valor (BRL)", "% do total"], [30, 18, 18, 12])

    produtos = ("Folha de Pagamento", "PIX", "Pagamento Fornecedores", "TM5")
    first = r + 1
    total_row = first + len(produtos)
    for i, p in enumerate(produtos):
        rr = first + i
        ws.cell(row=rr, column=1, value=p)
        ws.cell(row=rr, column=2, value=f"=COUNTIF(Base!$D$2:$D${last},$A{rr})")
        ws.cell(row=rr, column=3, value=f"=SUMIF(Base!$D$2:$D${last},$A{rr},Base!$F$2:$F${last})")
        ws.cell(row=rr, column=4, value=f"=IFERROR(C{rr}/$C${total_row},0)")
        for c in range(1, 5):
            ws.cell(row=rr, column=c).font = BODY
            ws.cell(row=rr, column=c).border = BOX
        ws.cell(row=rr, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=rr, column=3).number_format = FMT
        ws.cell(row=rr, column=4).number_format = PCT

    ws.cell(row=total_row, column=2, value=f"=SUM(B{first}:B{total_row - 1})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C{first}:C{total_row - 1})")
    ws.cell(row=total_row, column=4, value=1)
    _linha_total(ws, total_row, 4, 1, "Total Geral")
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=3).number_format = FMT
    ws.cell(row=total_row, column=4).number_format = PCT

    ws.cell(row=total_row + 2, column=1,
            value="Valores negativos = saída de caixa, conforme export SAP. "
                  "Totais calculados por fórmula sobre a aba Base.").font = NOTE


def _aba_produto(wb: Workbook, produto: str, base: pd.DataFrame, janela: "config.Janela", last: int) -> None:
    sub = (
        base[base["Produto"] == produto]
        .groupby("Vendor", as_index=False)
        .agg(Qtd=("Documento", "count"), Valor=("Valor", "sum"))
        .sort_values("Valor")
        .reset_index(drop=True)
    )
    ws = wb.create_sheet(produto[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"] = produto
    ws["A1"].font = TITLE
    ws["A2"] = f"Posting Date {janela.rotulo} · Status Pendentes"
    ws["A2"].font = NOTE
    _cabecalho(ws, 4, ["Fornecedor", "Qtd. documentos", "Valor (BRL)"], [52, 18, 18])

    f0 = 5
    for i, row in enumerate(sub.itertuples(index=False)):
        rr = f0 + i
        ws.cell(row=rr, column=1, value=row.Vendor)
        ws.cell(row=rr, column=2, value=row.Qtd)
        ws.cell(
            row=rr, column=3,
            value=f'=SUMIFS(Base!$F$2:$F${last},Base!$D$2:$D${last},"{produto}",'
                  f"Base!$E$2:$E${last},$A{rr})",
        )
        for c in range(1, 4):
            ws.cell(row=rr, column=c).font = BODY
            ws.cell(row=rr, column=c).border = BOX
        ws.cell(row=rr, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=rr, column=3).number_format = FMT

    tt = f0 + len(sub)
    ws.cell(row=tt, column=2, value=f"=SUM(B{f0}:B{tt - 1})")
    ws.cell(row=tt, column=3, value=f"=SUM(C{f0}:C{tt - 1})")
    _linha_total(ws, tt, 3, 1, "Total Geral")
    ws.cell(row=tt, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=tt, column=3).number_format = FMT
    ws.freeze_panes = "A5"


def _aba_validacoes(wb: Workbook, checks: list[tuple[str, str, str]], atencao: pd.DataFrame) -> None:
    ws = wb.create_sheet("Validações")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Log de validação"
    ws["A1"].font = TITLE
    _cabecalho(ws, 3, ["Check", "Severidade", "Detalhe"], [38, 14, 70])

    r = 4
    for check, sev, detalhe in checks:
        ws.cell(row=r, column=1, value=check).font = BODY
        sc = ws.cell(row=r, column=2, value=sev)
        sc.font, sc.fill = BOLD, _FILL_POR_SEV[sev]
        sc.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=detalhe).font = BODY
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = BOX
        r += 1

    r += 2
    ws.cell(row=r, column=1, value=f"Documentos ≥ {config.LIMITE_ATENCAO:,.2f} (revisão manual)").font = BOLD
    r += 1
    _cabecalho(ws, r, ["Documento", "Produto", "Fornecedor", "Valor (BRL)"], [38, 14, 70, 18])
    r += 1
    for row in atencao.itertuples(index=False):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font, cell.border = BODY, BOX
            if c == 4:
                cell.number_format = FMT
        r += 1


def escrever_relatorio(base: pd.DataFrame, janela: "config.Janela", destino: Path) -> Path:
    """Layout de referência: Base + Resumo (fórmulas) + uma aba por
    produto (fórmulas) + Validações. Fonte Arial 9, sem grade em todas
    as abas."""
    wb = Workbook()
    last = _aba_base(wb, base)
    _aba_resumo(wb, janela, base, last)
    for produto in ("Folha de Pagamento", "PIX", "Pagamento Fornecedores", "TM5"):
        _aba_produto(wb, produto, base, janela, last)

    checks = _checks(base, janela)
    atencao = (
        base[base["Valor"].abs() >= config.LIMITE_ATENCAO]
        .sort_values("Valor")
        .rename(columns={"Vendor": "Fornecedor"})
        .loc[:, ["Documento", "Produto", "Fornecedor", "Valor"]]
        .reset_index(drop=True)
    )
    _aba_validacoes(wb, checks, atencao)

    destino.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(destino)
    except PermissionError:
        # Mesmo nome de arquivo toda vez que roda no mesmo dia (janela.rotulo
        # não muda) — se já tiver um Payments_<data>.xlsx aberto no Excel de
        # uma rodada anterior, o Windows bloqueia a escrita. Em vez de
        # derrubar a rodada inteira, salva com um sufixo de horário.
        alternativo = destino.with_stem(f"{destino.stem}_{datetime.now():%H%M%S}")
        print(
            f"\nAVISO: não consegui salvar em {destino.name} — provavelmente está "
            f"aberto no Excel. Salvando como {alternativo.name} em vez disso."
        )
        wb.save(alternativo)
        return alternativo
    return destino


def _escrever_base_completa(wb: Workbook, completa: pd.DataFrame) -> None:
    ws = wb.create_sheet("Base")
    ws.sheet_view.showGridLines = False
    cols = list(completa.columns)
    larguras = []
    for nome in cols:
        if nome in ("Nome do fornecedor", "Texto de item"):
            larguras.append(40)
        elif nome in ("Produto", "RegraAplicada", "Banco"):
            larguras.append(24)
        else:
            larguras.append(max(12, len(nome) + 2))
    _cabecalho(ws, 1, cols, larguras)

    campos_data = {"Dt.lçto.cont.", "Data vencimento líq.", "Data de compensação", "Dt.lançamento"}
    campos_valor = {"Montante (ME)"}
    for r, row in enumerate(completa.itertuples(index=False), start=2):
        for cidx, (nome, v) in enumerate(zip(cols, row), start=1):
            cell = ws.cell(row=r, column=cidx, value=v)
            cell.font, cell.border = BODY, BOX
            if nome in campos_data and v is not None:
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = Alignment(horizontal="center")
            if nome in campos_valor:
                cell.number_format = FMT
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(completa)+1}"


def escrever_xlsx(limpo: pd.DataFrame, completa: pd.DataFrame, janela: "config.Janela", destino: Path) -> Path:
    base = limpo  # nome curto usado no resto da função (Resumo/SAP)
    wb = Workbook()

    ws = wb.active
    ws.title = "SAP"
    cols = ["Company Code", "G/L Account", "Documento", "Tipo Lançamento",
            "Vendor", "Produto", "Banco", "Status", "Posting Date", "Valor"]
    _cabecalho(ws, 1, cols, [16, 30, 14, 26, 42, 24, 16, 12, 14, 16])
    for r, row in enumerate(base[cols].itertuples(index=False), start=2):
        for cidx, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=cidx, value=v)
            cell.font, cell.border = BODY, BOX
            if cols[cidx - 1] == "Posting Date":
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = Alignment(horizontal="center")
            if cols[cidx - 1] == "Valor":
                cell.number_format = FMT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(base)+1}"
    ws.sheet_view.showGridLines = False

    resumo = wb.create_sheet("Resumo", 0)
    resumo.sheet_view.showGridLines = False
    resumo["B2"] = "G/L Account"
    resumo["C2"] = config.CONTA_GL
    resumo["B2"].font = BOLD
    _cabecalho(resumo, 4, ["Posting Date", "Status", "Produto", "Soma de Valor"], [14, 12, 26, 18], col_inicio=2)
    pivot = base.groupby("Produto", as_index=False)["Valor"].sum().sort_values("Valor")
    r = 5
    for row in pivot.itertuples(index=False):
        resumo.cell(row=r, column=2, value=janela.data_referencia).number_format = "DD/MM/YYYY"
        resumo.cell(row=r, column=3, value="Pendentes")
        resumo.cell(row=r, column=4, value=row.Produto)
        c5 = resumo.cell(row=r, column=5, value=round(row.Valor, 2))
        c5.number_format = FMT
        for col in range(2, 6):
            resumo.cell(row=r, column=col).font = BODY
        r += 1
    resumo.cell(row=r, column=3, value="Total Geral").font = BOLD
    tot = resumo.cell(row=r, column=5, value=round(pivot["Valor"].sum(), 2))
    tot.font, tot.number_format = BOLD, FMT

    # --- Dinâmica por produto (quebrada por banco) ---
    r += 3
    resumo.cell(row=r, column=2, value="Dinâmica por produto").font = BOLD
    r += 1
    linha_titulo = r
    titulos = ["Produto", "Qtd. documentos", "Santander", "Banco do Brasil", "Total (BRL)", "% do total"]
    larguras = [24, 16, 18, 18, 16, 12]
    for i, (nome, w) in enumerate(zip(titulos, larguras), start=2):
        cell = resumo.cell(row=linha_titulo, column=i, value=nome)
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if nome == "Santander":
            cell.fill, cell.font = SANTANDER_FILL, SANTANDER_FONT
        elif nome == "Banco do Brasil":
            cell.fill, cell.font = BB_FILL, BB_FONT
        else:
            cell.fill, cell.font = H_FILL, H_FONT
        resumo.column_dimensions[get_column_letter(i)].width = w

    dinamica = (
        base.pivot_table(index="Produto", columns="Banco", values="Valor", aggfunc="sum", fill_value=0.0)
        .reindex(columns=["Santander", "Banco do Brasil"], fill_value=0.0)
    )
    qtd = base.groupby("Produto").size()
    total_geral = base["Valor"].sum()

    r = linha_titulo + 1
    for produto in dinamica.index:
        santander = dinamica.loc[produto, "Santander"]
        bb = dinamica.loc[produto, "Banco do Brasil"]
        total_produto = santander + bb
        resumo.cell(row=r, column=2, value=produto).font = BODY
        resumo.cell(row=r, column=3, value=int(qtd[produto])).font = BODY
        c4 = resumo.cell(row=r, column=4, value=round(santander, 2))
        c4.font, c4.number_format = BODY, FMT
        c5 = resumo.cell(row=r, column=5, value=round(bb, 2))
        c5.font, c5.number_format = BODY, FMT
        c6 = resumo.cell(row=r, column=6, value=round(total_produto, 2))
        c6.font, c6.number_format = BODY, FMT
        c7 = resumo.cell(row=r, column=7, value=(total_produto / total_geral if total_geral else 0))
        c7.font, c7.number_format = BODY, PCT
        for col in range(2, 8):
            resumo.cell(row=r, column=col).border = BOX
        r += 1

    for col, valor in zip(
        range(2, 8),
        ["Total Geral", int(qtd.sum()), round(dinamica["Santander"].sum(), 2),
         round(dinamica["Banco do Brasil"].sum(), 2), round(total_geral, 2), 1.0],
    ):
        cell = resumo.cell(row=r, column=col, value=valor)
        cell.font, cell.fill, cell.border = BOLD, T_FILL, BOX
        if col in (4, 5, 6):
            cell.number_format = FMT
        if col == 7:
            cell.number_format = PCT

    _escrever_base_completa(wb, completa)

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera o relatório Payments a partir de dois exports SAP "
                     "('Administrar itens de fornecedor' e 'Partidas individuais no Razão')."
    )
    parser.add_argument(
        "--file1", type=Path, default=None,
        help="Caminho do primeiro arquivo XLSX de entrada. Se omitido junto com --file2, "
             "cai no comportamento antigo de escanear config.DIR_ENTRADA.",
    )
    parser.add_argument(
        "--file2", type=Path, default=None,
        help="Caminho do segundo arquivo XLSX de entrada.",
    )
    parser.add_argument(
        "--output-dir", type=Path, default=None,
        help="Pasta onde a subpasta <AAAA-MM-DD>/Payments_<data>.xlsx será criada. "
             "Default: config.DIR_SAIDA.",
    )
    return parser.parse_args(argv)


def _data_analisada(base: pd.DataFrame, fallback):
    """Data de fato encontrada nos lançamentos analisados (coluna Posting
    Date) — usada só pra nomear o arquivo/pasta de saída, pra não sair um
    Payments_<data errada>.xlsx quando alguém esquece de atualizar
    config.DATA_REFERENCIA na semana certa. O check 'Posting Date ==
    Data_Referencia' na aba Validações continua comparando contra o valor
    configurado (fallback aqui), então essa divergência não passa
    despercebida — só deixa de aparecer no NOME do arquivo."""
    datas = base["Posting Date"].dropna()
    if datas.empty:
        return fallback
    return datas.mode().iloc[0]


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    janela = config.janela_atual()

    if args.file1 and args.file2:
        df_banco, df_partidas = carregar_entrada_arquivos([args.file1, args.file2])
    else:
        df_banco, df_partidas = carregar_entrada(config.DIR_ENTRADA)
    print(f"Partidas individuais no Razão:   {len(df_banco)} linhas")
    print(f"Administrar itens de fornecedor: {len(df_partidas)} linhas")

    base, completa = classificar_semana(df_banco, df_partidas)

    print()
    resumo = base.groupby("Produto")["Valor"].sum().round(2)
    for produto, valor in resumo.items():
        print(f"  {produto:<25} {valor:>16,.2f}")
    print(f"  {'Total Geral':<25} {round(resumo.sum(),2):>16,.2f}")

    manuais = base[base["RegraAplicada"].str.contains("R6", na=False)]
    if not manuais.empty:
        print(f"\n{len(manuais)} documento(s) usando exceção manual (OVERRIDES_SEMANA):")
        for row in manuais.itertuples(index=False):
            print(f"  {row.Documento} | {row.Produto} | {row.Valor:,.2f}")
    else:
        print("\nNenhuma exceção manual usada — conferir se OVERRIDES_SEMANA está vazio de propósito.")

    saida_dir = args.output_dir if args.output_dir is not None else config.DIR_SAIDA
    data_analise = _data_analisada(base, janela.data_referencia)
    destino = (
        saida_dir
        / data_analise.strftime("%Y-%m-%d")
        / f"Payments_{data_analise.strftime('%d.%m.%Y')}.xlsx"
    )
    caminho = escrever_relatorio(base, janela, destino)
    print(f"\nArquivo: {caminho}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
